
from __future__ import annotations

import argparse
import hashlib
import json
import os
import smtplib
import webbrowser
from dataclasses import dataclass, field
from datetime import datetime
from email.message import EmailMessage
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# =========================================================
# DEFAULT CONFIG
# =========================================================
DEFAULT_CONFIG = {
    "output_dir": "output",
    "state_dir": ".ad_hoc_state",
    "retention_days": 365,
    "report_name_prefix": "ad_hoc_report",
    "default_source_type": "excel",
    # Use forward slashes on Windows to avoid unicode path issues
    "default_source_path": "ad_hoc_test.xlsx",
    "recipient_emails": ["bradleyzackery26@yahoo.com"],
    "email": {
        "send_email": False,
        "smtp_host": os.getenv("SMTP_HOST", "smtp.mail.yahoo.com"),
        "smtp_port": int(os.getenv("SMTP_PORT", "465")),
        "smtp_user": os.getenv("SMTP_USER", ""),
        "smtp_password": os.getenv("SMTP_PASSWORD", ""),
        "sender": os.getenv("SMTP_SENDER", os.getenv("SMTP_USER", "")),
        "subject_template": "Ad Hoc Report - {report_date}",
        "body_template": (
            "Hello,\n\n"
            "Attached is the latest ad hoc report generated on {report_date}.\n\n"
            "Regards,\n"
            "Automated Reporting Script"
        ),
    },
    "schema": {
        "date": ["date"],
        "units": ["amount_sold", "amount", "units", "qty", "quantity"],
        "unit_price": ["price", "unit_price", "price_per_unit", "rate"],
        "revenue": [
            "dollar_amount_in_sales",
            "sales",
            "revenue",
            "total_sales",
            "amount_total",
            "total_amount",
        ],
    },
}


# =========================================================
# UTILITIES
# =========================================================
def ensure_dir(path: Path) -> Path:
    path.mkdir(parents=True, exist_ok=True)
    return path


def normalize_col(col: str) -> str:
    return str(col).strip().lower().replace(" ", "_")


def parse_date_series(series: pd.Series) -> pd.Series:
    return pd.to_datetime(series, errors="coerce")


def row_hash(df: pd.DataFrame) -> pd.Series:
    canonical = df.fillna("").astype(str).apply(lambda row: "|".join(row.values), axis=1)
    return canonical.apply(lambda x: hashlib.sha256(x.encode("utf-8")).hexdigest())


def resolve_source_path(source_path: str, script_dir: Path) -> Path:
    sp = Path(source_path)
    if sp.is_absolute():
        return sp
    candidate = script_dir / sp
    if candidate.exists():
        return candidate
    return sp


def autosize_worksheet(ws, max_width: int = 30) -> None:
    for col_cells in ws.columns:
        try:
            max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col_cells)
        except TypeError:
            continue
        adjusted = min(max(max_len + 2, 10), max_width)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = adjusted


# =========================================================
# DATA SOURCE LOADING
# =========================================================
@dataclass
class LoadedSource:
    source_name: str
    sheets: Dict[str, pd.DataFrame] = field(default_factory=dict)


def load_excel_source(path: str) -> LoadedSource:
    sheets = pd.read_excel(path, sheet_name=None, engine="openpyxl")
    cleaned = {
        sheet_name: df.copy()
        for sheet_name, df in sheets.items()
        if df is not None and not df.empty
    }
    return LoadedSource(source_name=Path(path).stem, sheets=cleaned)


def load_google_sheets_source(sheet_url: str, service_account_json: Optional[str] = None) -> LoadedSource:
    try:
        import gspread
    except ImportError as exc:
        raise ImportError(
            "Google Sheets support requires gspread and google-auth. "
            "Install with: pip install gspread google-auth"
        ) from exc

    gc = gspread.service_account(filename=service_account_json) if service_account_json else gspread.service_account()
    sh = gc.open_by_url(sheet_url)

    sheets = {}
    for ws in sh.worksheets():
        records = ws.get_all_records()
        if records:
            sheets[ws.title] = pd.DataFrame(records)

    return LoadedSource(source_name=sh.title, sheets=sheets)


# =========================================================
# STANDARDIZATION
# =========================================================
def pick_matching_column(columns: List[str], aliases: List[str]) -> Optional[str]:
    normalized_columns = [normalize_col(c) for c in columns]
    normalized_aliases = [normalize_col(a) for a in aliases]
    for alias in normalized_aliases:
        if alias in normalized_columns:
            return alias
    return None


def standardize_sheet(df: pd.DataFrame, sheet_name: str, schema_cfg: dict) -> pd.DataFrame:
    tmp = df.copy()
    original_cols = list(tmp.columns)
    tmp.columns = [normalize_col(c) for c in tmp.columns]

    date_col = pick_matching_column(tmp.columns.tolist(), schema_cfg.get("date", []))
    units_col = pick_matching_column(tmp.columns.tolist(), schema_cfg.get("units", []))
    unit_price_col = pick_matching_column(tmp.columns.tolist(), schema_cfg.get("unit_price", []))
    revenue_col = pick_matching_column(tmp.columns.tolist(), schema_cfg.get("revenue", []))

    out = pd.DataFrame()
    out["source_sheet"] = [sheet_name] * len(tmp)
    out["date"] = parse_date_series(tmp[date_col]) if date_col else pd.NaT
    out["units"] = pd.to_numeric(tmp[units_col], errors="coerce") if units_col else pd.NA
    out["unit_price"] = pd.to_numeric(tmp[unit_price_col], errors="coerce") if unit_price_col else pd.NA
    out["revenue"] = pd.to_numeric(tmp[revenue_col], errors="coerce") if revenue_col else pd.NA

    if out["revenue"].isna().all() and out["units"].notna().any() and out["unit_price"].notna().any():
        out["revenue"] = out["units"] * out["unit_price"]

    valid_units = out["units"].fillna(0) != 0
    missing_price = out["unit_price"].isna()
    can_backfill = valid_units & missing_price & out["revenue"].notna()
    out.loc[can_backfill, "unit_price"] = out.loc[can_backfill, "revenue"] / out.loc[can_backfill, "units"]

    for col in original_cols:
        ncol = normalize_col(col)
        out[f"raw_{ncol}"] = tmp[ncol]

    out["year_month"] = out["date"].dt.to_period("M").astype(str)
    out["load_timestamp"] = pd.Timestamp.now().floor("s")
    return out


def combine_and_standardize(loaded: LoadedSource, config: dict) -> pd.DataFrame:
    frames = []
    for sheet_name, df in loaded.sheets.items():
        frames.append(standardize_sheet(df, sheet_name, config["schema"]))

    if not frames:
        return pd.DataFrame(
            columns=[
                "source_sheet",
                "date",
                "units",
                "unit_price",
                "revenue",
                "year_month",
                "load_timestamp",
            ]
        )

    combined = pd.concat(frames, ignore_index=True)
    combined = combined.sort_values(["date", "source_sheet"], na_position="last").reset_index(drop=True)

    dedupe_cols = [c for c in ["source_sheet", "date", "units", "unit_price", "revenue"] if c in combined.columns]
    combined = combined.drop_duplicates(subset=dedupe_cols, keep="last").reset_index(drop=True)

    return combined


# =========================================================
# STATE / RETENTION
# =========================================================
def load_state_hashes(state_dir: Path) -> pd.DataFrame:
    path = state_dir / "processed_row_hashes.csv"
    if path.exists():
        return pd.read_csv(path)
    return pd.DataFrame(columns=["row_hash", "first_seen"])


def save_state_hashes(state_dir: Path, state_df: pd.DataFrame, retention_days: int) -> None:
    cutoff = pd.Timestamp.now().normalize() - pd.Timedelta(days=retention_days)

    if not state_df.empty:
        state_df["first_seen"] = pd.to_datetime(state_df["first_seen"], errors="coerce")
        state_df = state_df[state_df["first_seen"] >= cutoff].copy()
        state_df["first_seen"] = state_df["first_seen"].dt.strftime("%Y-%m-%d %H:%M:%S")

    state_df.to_csv(state_dir / "processed_row_hashes.csv", index=False)


def detect_new_rows(df: pd.DataFrame, state_dir: Path, retention_days: int) -> Tuple[pd.DataFrame, pd.DataFrame]:
    df = df.copy()
    hash_input_cols = [c for c in df.columns if c != "load_timestamp"]
    df["row_hash"] = row_hash(df[hash_input_cols])

    state_df = load_state_hashes(state_dir)
    seen_hashes = set(state_df["row_hash"].astype(str).tolist()) if not state_df.empty else set()

    new_rows_df = df.loc[~df["row_hash"].isin(seen_hashes)].copy()

    now = pd.Timestamp.now().floor("s")
    new_hashes = pd.DataFrame({
        "row_hash": new_rows_df["row_hash"].tolist(),
        "first_seen": [now] * len(new_rows_df)
    })

    updated_state = pd.concat([state_df, new_hashes], ignore_index=True).drop_duplicates(
        subset=["row_hash"], keep="first"
    )
    save_state_hashes(state_dir, updated_state, retention_days)

    return df, new_rows_df


def append_history(df: pd.DataFrame, state_dir: Path, retention_days: int) -> pd.DataFrame:
    history_path = state_dir / "data_history.csv"
    historical = pd.read_csv(history_path) if history_path.exists() else pd.DataFrame(columns=df.columns)

    combined = pd.concat([historical, df], ignore_index=True)

    if "date" in combined.columns:
        combined["date"] = pd.to_datetime(combined["date"], errors="coerce")
        cutoff = pd.Timestamp.now().normalize() - pd.Timedelta(days=retention_days)
        combined = combined[(combined["date"].isna()) | (combined["date"] >= cutoff)]

    dedupe_cols = [c for c in ["source_sheet", "date", "units", "unit_price", "revenue", "row_hash"] if c in combined.columns]
    combined = combined.drop_duplicates(subset=dedupe_cols, keep="last")

    save_copy = combined.copy()
    if "date" in save_copy.columns:
        save_copy["date"] = save_copy["date"].dt.strftime("%Y-%m-%d")

    save_copy.to_csv(history_path, index=False)
    return combined


# =========================================================
# REPORTING CALCULATIONS
# =========================================================
def build_summary_tables(history_df: pd.DataFrame, new_rows_df: pd.DataFrame) -> Dict[str, pd.DataFrame]:
    df = history_df.copy()

    if not df.empty and "date" in df.columns:
        df["date"] = pd.to_datetime(df["date"], errors="coerce")

    total_revenue = float(df["revenue"].fillna(0).sum()) if "revenue" in df.columns and not df.empty else 0.0
    total_units = float(df["units"].fillna(0).sum()) if "units" in df.columns and not df.empty else 0.0
    rows_loaded = int(len(df))
    new_rows = int(len(new_rows_df))
    avg_daily_revenue = (
        float(df.groupby(df["date"].dt.date)["revenue"].sum().mean())
        if not df.empty and df["date"].notna().any()
        else 0.0
    )
    avg_unit_price = (
        float(pd.to_numeric(df["unit_price"], errors="coerce").fillna(0).mean())
        if "unit_price" in df.columns and not df.empty
        else 0.0
    )

    metrics = pd.DataFrame({
        "metric": [
            "rows_loaded",
            "new_rows",
            "total_revenue",
            "total_units",
            "avg_daily_revenue",
            "avg_unit_price",
        ],
        "value": [
            rows_loaded,
            new_rows,
            total_revenue,
            total_units,
            avg_daily_revenue,
            avg_unit_price,
        ],
    })

    if df.empty or not df["date"].notna().any():
        return {
            "metrics": metrics,
            "daily": pd.DataFrame(),
            "by_sheet": pd.DataFrame(),
            "monthly": pd.DataFrame(),
        }

    daily = (
        df.groupby(df["date"].dt.date, dropna=True)
        .agg(
            total_revenue=("revenue", "sum"),
            total_units=("units", "sum"),
            rows=("date", "size"),
        )
        .reset_index()
    )
    daily.columns = ["day", "total_revenue", "total_units", "rows"]
    daily["day"] = pd.to_datetime(daily["day"]).astype(str)

    by_sheet = (
        df.groupby("source_sheet", dropna=False)
        .agg(
            total_revenue=("revenue", "sum"),
            total_units=("units", "sum"),
            rows=("source_sheet", "size"),
        )
        .reset_index()
        .sort_values("total_revenue", ascending=False)
    )

    monthly = (
        df.assign(month=df["date"].dt.to_period("M").astype(str))
        .groupby("month", dropna=False)
        .agg(
            total_revenue=("revenue", "sum"),
            total_units=("units", "sum"),
            rows=("month", "size"),
        )
        .reset_index()
        .sort_values("month")
    )

    return {
        "metrics": metrics,
        "daily": daily,
        "by_sheet": by_sheet,
        "monthly": monthly,
    }


# =========================================================
# VISUALS
# =========================================================
def create_chart_pngs(summary: Dict[str, pd.DataFrame], chart_dir: Path) -> List[Path]:
    ensure_dir(chart_dir)
    chart_paths: List[Path] = []

    daily = summary.get("daily", pd.DataFrame())
    if not daily.empty:
        p = chart_dir / "daily_revenue_trend.png"
        fig, ax = plt.subplots(figsize=(10, 4.5))
        ax.plot(pd.to_datetime(daily["day"]), daily["total_revenue"], marker="o")
        ax.set_title("Daily Revenue Trend")
        ax.set_xlabel("Date")
        ax.set_ylabel("Revenue")
        ax.grid(True, alpha=0.3)
        fig.autofmt_xdate()
        fig.tight_layout()
        fig.savefig(p, dpi=150)
        plt.close(fig)
        chart_paths.append(p)

    by_sheet = summary.get("by_sheet", pd.DataFrame())
    if not by_sheet.empty:
        p = chart_dir / "revenue_by_sheet.png"
        fig, ax = plt.subplots(figsize=(8, 4.5))
        ax.bar(by_sheet["source_sheet"].astype(str), by_sheet["total_revenue"])
        ax.set_title("Revenue by Source Sheet")
        ax.set_xlabel("Source Sheet")
        ax.set_ylabel("Revenue")
        plt.xticks(rotation=15)
        fig.tight_layout()
        fig.savefig(p, dpi=150)
        plt.close(fig)
        chart_paths.append(p)

    monthly = summary.get("monthly", pd.DataFrame())
    if not monthly.empty:
        p = chart_dir / "monthly_units_and_revenue.png"
        fig, ax1 = plt.subplots(figsize=(10, 4.5))
        ax1.bar(monthly["month"].astype(str), monthly["total_units"], alpha=0.7)
        ax1.set_title("Monthly Units and Revenue")
        ax1.set_xlabel("Month")
        ax1.set_ylabel("Units")

        ax2 = ax1.twinx()
        ax2.plot(monthly["month"].astype(str), monthly["total_revenue"], color="tab:red", marker="o")
        ax2.set_ylabel("Revenue")

        fig.tight_layout()
        fig.savefig(p, dpi=150)
        plt.close(fig)
        chart_paths.append(p)

    return chart_paths


# =========================================================
# EXCEL OUTPUT STYLING
# =========================================================
THIN_GRAY = Side(style="thin", color="D9D9D9")
CARD_BORDER = Border(left=THIN_GRAY, right=THIN_GRAY, top=THIN_GRAY, bottom=THIN_GRAY)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center")

METRIC_STYLE_MAP = {
    "rows_loaded": {"fill": "D9EAF7", "font": "0B3C5D", "fmt": "0"},
    "new_rows": {"fill": "E8DAEF", "font": "5B2C6F", "fmt": "0"},
    "total_revenue": {"fill": "D5F5E3", "font": "186A3B", "fmt": '$#,##0.00'},
    "total_units": {"fill": "FCF3CF", "font": "7D6608", "fmt": '#,##0'},
    "avg_daily_revenue": {"fill": "FADBD8", "font": "922B21", "fmt": '$#,##0.00'},
    "avg_unit_price": {"fill": "D6EAF8", "font": "1B4F72", "fmt": '$#,##0.00'},
}


def style_summary_metrics(ws) -> None:
    """Color-code the metrics table on Ad_Hoc_Report sheet."""
    ws["A1"] = "Summary Metrics"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"].font = Font(bold=True)
    ws["B2"].font = Font(bold=True)
    ws["A2"].alignment = CENTER
    ws["B2"].alignment = CENTER
    ws["A2"].fill = PatternFill("solid", fgColor="1F4E78")
    ws["B2"].fill = PatternFill("solid", fgColor="1F4E78")
    ws["A2"].font = Font(color="FFFFFF", bold=True)
    ws["B2"].font = Font(color="FFFFFF", bold=True)

    row = 3
    while True:
        metric_cell = ws[f"A{row}"]
        value_cell = ws[f"B{row}"]
        metric = metric_cell.value
        if metric is None:
            break
        metric_key = str(metric).strip()
        style = METRIC_STYLE_MAP.get(metric_key, {"fill": "F2F2F2", "font": "000000", "fmt": 'General'})

        for cell in [metric_cell, value_cell]:
            cell.fill = PatternFill("solid", fgColor=style["fill"])
            cell.font = Font(color=style["font"], bold=True)
            cell.alignment = CENTER
            cell.border = CARD_BORDER
        value_cell.number_format = style["fmt"]
        row += 1


def add_kpi_card(ws, start_row: int, start_col: int, title: str, value, fill_color: str, font_color: str, number_format: str) -> None:
    """Draw a KPI card using merged cells and formatting."""
    # 4 columns wide x 4 rows high
    end_col = start_col + 3
    title_row = start_row
    value_row_start = start_row + 1
    value_row_end = start_row + 3

    ws.merge_cells(start_row=title_row, start_column=start_col, end_row=title_row, end_column=end_col)
    ws.merge_cells(start_row=value_row_start, start_column=start_col, end_row=value_row_end, end_column=end_col)

    title_cell = ws.cell(title_row, start_col)
    value_cell = ws.cell(value_row_start, start_col)

    title_cell.value = title
    value_cell.value = value
    value_cell.number_format = number_format

    title_fill = PatternFill("solid", fgColor=fill_color)
    value_fill = PatternFill("solid", fgColor=fill_color)

    title_cell.fill = title_fill
    value_cell.fill = value_fill

    title_cell.font = Font(color="FFFFFF", bold=True, size=10)
    value_cell.font = Font(color=font_color, bold=True, size=18)

    title_cell.alignment = CENTER
    value_cell.alignment = CENTER

    # apply fill/border to all cells in card region
    for r in range(start_row, start_row + 4):
        ws.row_dimensions[r].height = 24 if r == start_row else 26
        for c in range(start_col, end_col + 1):
            cell = ws.cell(r, c)
            cell.fill = title_fill if r == start_row else value_fill
            cell.border = CARD_BORDER
            if r != start_row and c != start_col:
                # merged helper cells
                cell.alignment = CENTER


def build_visuals_dashboard(ws, summary: Dict[str, pd.DataFrame], chart_paths: List[Path]) -> None:
    """Create KPI cards + grid-based chart layout on Visuals sheet."""
    # Title area
    ws["A1"] = "Ad Hoc Report Dashboard"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A2"] = "KPI cards and visuals are arranged for readability."
    ws["A2"].font = Font(italic=True, size=10)

    # Wider columns for dashboard spacing
    for col in ["A", "B", "C", "D", "E", "F", "G", "H", "I",
                "J", "K", "L", "M", "N", "O", "P", "Q", "R"]:
        ws.column_dimensions[col].width = 14

    # KPI values from metrics table
    metrics_df = summary.get("metrics", pd.DataFrame())
    metric_map = {}
    if not metrics_df.empty:
        metric_map = dict(zip(metrics_df["metric"], metrics_df["value"]))

    # 6 KPI cards in a 3x2 layout
    cards = [
        (4, 1,  "Rows Loaded",       metric_map.get("rows_loaded", 0),       "5B9BD5", "FFFFFF", '0'),
        (4, 6,  "New Rows",          metric_map.get("new_rows", 0),          "8064A2", "FFFFFF", '0'),
        (4, 11, "Total Revenue",     metric_map.get("total_revenue", 0),     "70AD47", "FFFFFF", '$#,##0.00'),
        (9, 1,  "Total Units",       metric_map.get("total_units", 0),       "ED7D31", "FFFFFF", '#,##0'),
        (9, 6,  "Avg Daily Revenue", metric_map.get("avg_daily_revenue", 0), "C55A11", "FFFFFF", '$#,##0.00'),
        (9, 11, "Avg Unit Price",    metric_map.get("avg_unit_price", 0),    "4472C4", "FFFFFF", '$#,##0.00'),
    ]
    for r, c, title, value, fill, font_color, num_fmt in cards:
        add_kpi_card(ws, r, c, title, value, fill, font_color, num_fmt)

    # Chart positions (start lower so KPI cards have room)
    chart_anchors = ["A16", "J16", "A40", "J40", "A64", "J64"]
    title_cells = ["A15", "J15", "A39", "J39", "A63", "J63"]
    default_titles = [
        "Daily Revenue Trend",
        "Revenue by Source Sheet",
        "Monthly Units and Revenue",
        "Chart 4",
        "Chart 5",
        "Chart 6",
    ]

    target_width = 540
    target_height = 280

    for idx, chart_path in enumerate(chart_paths):
        if idx >= len(chart_anchors):
            break
        if chart_path.exists():
            ws[title_cells[idx]] = default_titles[idx]
            ws[title_cells[idx]].font = Font(bold=True, size=11)
            img = XLImage(str(chart_path))
            img.width = target_width
            img.height = target_height
            ws.add_image(img, chart_anchors[idx])

    for row_num in [15, 16, 39, 40, 63, 64]:
        ws.row_dimensions[row_num].height = 22


def write_report_workbook(
    output_path: Path,
    history_df: pd.DataFrame,
    new_rows_df: pd.DataFrame,
    summary: Dict[str, pd.DataFrame],
    chart_paths: List[Path],
    meta: dict,
) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        history_df.to_excel(writer, sheet_name="Normalized_Data", index=False)
        new_rows_df.to_excel(writer, sheet_name="New_Rows", index=False)
        summary["metrics"].to_excel(writer, sheet_name="Ad_Hoc_Report", index=False, startrow=1)

        current_row = len(summary["metrics"]) + 5
        if not summary["by_sheet"].empty:
            summary["by_sheet"].to_excel(writer, sheet_name="Ad_Hoc_Report", index=False, startrow=current_row)
            current_row += len(summary["by_sheet"]) + 3

        if not summary["monthly"].empty:
            summary["monthly"].to_excel(writer, sheet_name="Ad_Hoc_Report", index=False, startrow=current_row)

        run_log = pd.DataFrame([
            {"field": "generated_at", "value": meta.get("generated_at")},
            {"field": "source_name", "value": meta.get("source_name")},
            {"field": "retention_days", "value": meta.get("retention_days")},
            {"field": "rows_in_history", "value": len(history_df)},
            {"field": "new_rows_in_run", "value": len(new_rows_df)},
            {"field": "report_file", "value": output_path.name},
        ])
        run_log.to_excel(writer, sheet_name="Run_Log", index=False)

        pd.DataFrame({"info": ["Dashboard visuals inserted below"]}).to_excel(writer, sheet_name="Visuals", index=False)

    wb = load_workbook(output_path)

    for sheet_name in ["Normalized_Data", "New_Rows", "Ad_Hoc_Report", "Run_Log", "Visuals"]:
        if sheet_name in wb.sheetnames:
            autosize_worksheet(wb[sheet_name], max_width=28)

    # Style the summary metrics table on Ad_Hoc_Report sheet
    if "Ad_Hoc_Report" in wb.sheetnames:
        style_summary_metrics(wb["Ad_Hoc_Report"])

    # Build KPI cards + chart dashboard on Visuals sheet
    if "Visuals" in wb.sheetnames:
        build_visuals_dashboard(wb["Visuals"], summary, chart_paths)

    wb.save(output_path)
    return output_path


# =========================================================
# EMAIL
# =========================================================
def send_email_with_attachment(report_path: Path, config: dict) -> None:
    email_cfg = config.get("email", {})
    recipients = config.get("recipient_emails", [])

    if not email_cfg.get("send_email", False):
        print("Email sending disabled by config. Report created but not emailed.")
        return

    smtp_host = email_cfg.get("smtp_host")
    smtp_port = int(email_cfg.get("smtp_port", 465))
    smtp_user = email_cfg.get("smtp_user")
    smtp_password = email_cfg.get("smtp_password")
    sender = email_cfg.get("sender") or smtp_user

    if not all([smtp_host, smtp_port, smtp_user, smtp_password, sender]):
        raise ValueError("SMTP settings are incomplete. Update your config before enabling email.")

    report_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    msg = EmailMessage()
    msg["Subject"] = email_cfg.get("subject_template", "Ad Hoc Report - {report_date}").format(report_date=report_date)
    msg["From"] = sender
    msg["To"] = ", ".join(recipients)
    msg.set_content(email_cfg.get("body_template", "Attached is the latest ad hoc report.").format(report_date=report_date))

    with open(report_path, "rb") as f:
        msg.add_attachment(
            f.read(),
            maintype="application",
            subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=report_path.name,
        )

    with smtplib.SMTP_SSL(smtp_host, smtp_port) as smtp:
        smtp.login(smtp_user, smtp_password)
        smtp.send_message(msg)

    print(f"Email sent to: {', '.join(recipients)}")


# =========================================================
# OUTPUT ACCESS HELPERS
# =========================================================
def create_clickable_report_link(report_path: Path) -> Path:
    report_path = report_path.resolve()
    html_path = report_path.with_name(f"open_{report_path.stem}.html")
    file_uri = report_path.as_uri()

    html = f"""<!DOCTYPE html>
<html>
<head>
    <meta charset=\"utf-8\">
    <meta http-equiv=\"refresh\" content=\"0; url={file_uri}\">
    <title>Open Report</title>
</head>
<body>
    <p>Your report is ready.</p>
    <p><a href=\"{file_uri}\">Click here to open the report</a></p>
    <p>Report location: {report_path}</p>
</body>
</html>
"""
    html_path.write_text(html, encoding="utf-8")
    return html_path


def open_output_location(report_path: Path) -> None:
    try:
        webbrowser.open(report_path.resolve().as_uri())
    except Exception:
        try:
            webbrowser.open(report_path.resolve().parent.as_uri())
        except Exception:
            pass


# =========================================================
# CONFIG + RUNNER
# =========================================================
def load_config(path: Optional[str] = None) -> dict:
    config = json.loads(json.dumps(DEFAULT_CONFIG))

    if path and Path(path).exists():
        with open(path, "r", encoding="utf-8") as f:
            user_cfg = json.load(f)

        for k, v in user_cfg.items():
            if isinstance(v, dict) and isinstance(config.get(k), dict):
                config[k].update(v)
            else:
                config[k] = v

    return config


def run(
    source_type: Optional[str] = None,
    source_path: Optional[str] = None,
    config_path: Optional[str] = None,
    google_credentials: Optional[str] = None,
) -> Path:
    script_dir = Path(__file__).resolve().parent
    config = load_config(config_path)

    source_type = source_type or config.get("default_source_type", "excel")
    source_path = source_path or config.get("default_source_path", "ad_hoc_test.xlsx")
    source_path_obj = resolve_source_path(source_path, script_dir)

    output_dir = ensure_dir(script_dir / config["output_dir"])
    state_dir = ensure_dir(script_dir / config["state_dir"])
    retention_days = int(config.get("retention_days", 365))

    if source_type == "excel":
        if not source_path_obj.exists():
            raise FileNotFoundError(
                f"Excel source file not found: {source_path_obj}\n"
                f"Tip: either place the file in {script_dir} or pass --source-path with the full file path."
            )
        loaded = load_excel_source(str(source_path_obj))
    elif source_type == "google_sheets":
        loaded = load_google_sheets_source(str(source_path_obj), google_credentials)
    else:
        raise ValueError("source_type must be 'excel' or 'google_sheets'")

    standardized = combine_and_standardize(loaded, config)
    standardized_with_hash, new_rows_df = detect_new_rows(standardized, state_dir, retention_days)
    history_df = append_history(standardized_with_hash, state_dir, retention_days)
    summaries = build_summary_tables(history_df, new_rows_df)

    chart_dir = ensure_dir(output_dir / "charts")
    chart_paths = create_chart_pngs(summaries, chart_dir)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    report_path = output_dir / f"{config['report_name_prefix']}_{timestamp}.xlsx"

    meta = {
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "source_name": loaded.source_name,
        "retention_days": retention_days,
    }

    write_report_workbook(report_path, history_df, new_rows_df, summaries, chart_paths, meta)
    send_email_with_attachment(report_path, config)
    return report_path


# =========================================================
# CLI
# =========================================================
def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Automated ad hoc report generator")
    parser.add_argument(
        "--source-type",
        choices=["excel", "google_sheets"],
        default=None,
        help="Defaults to config default_source_type if omitted",
    )
    parser.add_argument(
        "--source-path",
        default=None,
        help="Defaults to config default_source_path if omitted",
    )
    parser.add_argument(
        "--config",
        default=None,
        help="Optional path to JSON config file",
    )
    parser.add_argument(
        "--google-credentials",
        required=False,
        help="Path to Google service account JSON file",
    )
    return parser


def main() -> None:
    parser = build_arg_parser()
    args = parser.parse_args()

    report_path = run(
        source_type=args.source_type,
        source_path=args.source_path,
        config_path=args.config,
        google_credentials=args.google_credentials,
    )

    report_path = Path(report_path).resolve()
    link_file = create_clickable_report_link(report_path)

    print(f"Report created: {report_path}")
    print(f"Report link: {report_path.as_uri()}")
    print(f"Open-this-link file: {link_file}")
    print(f"Open-this-link URI: {link_file.resolve().as_uri()}")

    open_output_location(report_path)


if __name__ == "__main__":
    main()
