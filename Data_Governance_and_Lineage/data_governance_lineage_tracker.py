#!/usr/bin/env python3
"""
Data Governance & Lineage Tracker + Excel Packaging
--------------------------------------------------
A portfolio-ready Python script that simulates a lightweight data governance
and lineage framework for a multi-table e-commerce dataset (Olist-style), and
also packages all generated CSV artifacts plus the lineage graph into a single
Excel workbook.

What this version adds:
1) Automatically searches multiple folders for input CSV files
2) Supports BOTH simplified filenames and original Olist filenames
3) Prints helpful diagnostics before failing
4) Exports all result CSVs to an Excel workbook (one sheet per CSV)
5) Adds the lineage graph image to a dedicated Excel sheet, sized to be easy to read
   and positioned so it does not overlap worksheet content

Examples:
    python data_governance_lineage_tracker_with_excel.py
    python data_governance_lineage_tracker_with_excel.py --data-dir ./data
    python data_governance_lineage_tracker_with_excel.py --data-dir . --output-dir ./governance_output
"""

from __future__ import annotations

import argparse
from pathlib import Path
from typing import Dict, List, Tuple, Optional

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment

try:
    import networkx as nx
except Exception:
    nx = None

try:
    from PIL import Image as PILImage
except Exception:
    PILImage = None


# -----------------------------
# Configuration
# -----------------------------
FILE_MAP_VARIANTS = [
    {
        "customers": "customers.csv",
        "orders": "orders.csv",
        "order_items": "order_items.csv",
        "payments": "payments.csv",
        "products": "products.csv",
        "order_reviews": "order_reviews.csv",  # optional
    },
    {
        "customers": "olist_customers_dataset.csv",
        "orders": "olist_orders_dataset.csv",
        "order_items": "olist_order_items_dataset.csv",
        "payments": "olist_order_payments_dataset.csv",
        "products": "olist_products_dataset.csv",
        "order_reviews": "olist_order_reviews_dataset.csv",  # optional
    },
]

DOMAIN_MAP = {
    "customers": "Customer",
    "orders": "Sales",
    "order_items": "Sales",
    "payments": "Finance",
    "products": "Product",
    "order_reviews": "Customer Experience",
}

OWNER_MAP = {
    "customers": "Customer Data Steward",
    "orders": "Sales Data Steward",
    "order_items": "Sales Data Steward",
    "payments": "Finance Data Steward",
    "products": "Product Data Steward",
    "order_reviews": "CX Data Steward",
    "dim_customer": "Analytics Engineering",
    "dim_product": "Analytics Engineering",
    "fact_order": "Analytics Engineering",
    "sales_summary_daily": "BI / Analytics",
}

UPDATE_FREQUENCY = {
    "customers": "Daily",
    "orders": "Daily",
    "order_items": "Daily",
    "payments": "Daily",
    "products": "Weekly",
    "order_reviews": "Daily",
    "dim_customer": "Daily",
    "dim_product": "Daily",
    "fact_order": "Daily",
    "sales_summary_daily": "Daily",
}

PK_MAP = {
    "customers": ["customer_id"],
    "orders": ["order_id"],
    "order_items": ["order_id", "order_item_id"],
    "payments": ["order_id", "payment_sequential"],
    "products": ["product_id"],
    "order_reviews": ["review_id"],
    "dim_customer": ["customer_id"],
    "dim_product": ["product_id"],
    "fact_order": ["order_id", "order_item_id"],
    "sales_summary_daily": ["order_purchase_date"],
}

FK_MAP = {
    "orders": [("customer_id", "customers", "customer_id")],
    "order_items": [("order_id", "orders", "order_id"), ("product_id", "products", "product_id")],
    "payments": [("order_id", "orders", "order_id")],
    "order_reviews": [("order_id", "orders", "order_id")],
    "fact_order": [
        ("order_id", "orders", "order_id"),
        ("customer_id", "customers", "customer_id"),
        ("product_id", "products", "product_id"),
    ],
}

PII_PATTERNS = {
    "Direct Identifier": ["email", "phone", "name", "cpf", "ssn"],
    "Location": ["zip", "city", "state", "address"],
    "Behavioral": ["review", "comment"],
}

SENSITIVE_PATTERNS = ["payment", "credit", "installment", "value"]
DATE_HINTS = ["date", "timestamp", "dt"]
MEASURE_HINTS = ["price", "value", "amount", "qty", "quantity", "score", "weight", "cm", "g"]

# Fixed sheet names for readability and Excel compatibility
CSV_SHEET_NAMES = {
    "data_catalog.csv": "Data_Catalog",
    "column_catalog.csv": "Column_Catalog",
    "data_quality_report.csv": "Quality_Report",
    "lineage_edges.csv": "Lineage_Edges",
    "column_lineage.csv": "Column_Lineage",
    "process_registry.csv": "Process_Registry",
    "pii_summary.csv": "PII_Summary",
    "run_summary.csv": "Run_Summary",
    "dim_customer.csv": "Dim_Customer",
    "dim_product.csv": "Dim_Product",
    "fact_order.csv": "Fact_Order",
    "sales_summary_daily.csv": "Sales_Summary_Daily",
}


# -----------------------------
# Utility functions
# -----------------------------
def safe_read_csv(path: Path) -> pd.DataFrame:
    return pd.read_csv(path)


def dtype_category(series: pd.Series) -> str:
    if pd.api.types.is_bool_dtype(series):
        return "boolean"
    if pd.api.types.is_numeric_dtype(series):
        return "numeric"
    if pd.api.types.is_datetime64_any_dtype(series):
        return "datetime"
    return "string"


def classify_column(col_name: str) -> Tuple[str, str]:
    c = col_name.lower()
    for classification, patterns in PII_PATTERNS.items():
        if any(p in c for p in patterns):
            return ("Restricted", classification)
    if any(p in c for p in SENSITIVE_PATTERNS):
        return ("Confidential", "Financial")
    if c in {"order_id", "customer_id", "product_id"} or c.endswith("_id"):
        return ("Internal", "Business Key")
    if any(h in c for h in DATE_HINTS):
        return ("Internal", "Operational Timestamp")
    if any(h in c for h in MEASURE_HINTS):
        return ("Internal", "Measure")
    return ("Public", "Descriptive Attribute")


def detect_candidate_keys(df: pd.DataFrame) -> List[str]:
    candidate_keys = []
    n = len(df)
    if n == 0:
        return candidate_keys
    for col in df.columns:
        non_null = df[col].notna().sum()
        unique_non_null = df[col].nunique(dropna=True)
        if non_null == n and unique_non_null == n:
            candidate_keys.append(col)
    return candidate_keys


def contains_pii(df: pd.DataFrame) -> bool:
    return any(classify_column(c)[0] == "Restricted" for c in df.columns)


def infer_criticality(table_name: str) -> str:
    if table_name in {"orders", "order_items", "payments", "fact_order", "sales_summary_daily"}:
        return "High"
    if table_name in {"customers", "products", "dim_customer", "dim_product"}:
        return "Medium"
    return "Low"


def infer_retention(table_name: str) -> str:
    if table_name in {"payments", "fact_order", "sales_summary_daily"}:
        return "7 years"
    if table_name in {"customers", "dim_customer"}:
        return "3 years after last activity"
    return "2 years"


def infer_table_description(table_name: str) -> str:
    descriptions = {
        "customers": "Customer master data captured from the commerce platform.",
        "orders": "Order header records representing customer purchases.",
        "order_items": "Order line items including product and pricing details.",
        "payments": "Payment details and financial settlement attributes by order.",
        "products": "Product catalog and descriptive product attributes.",
        "order_reviews": "Customer review text and score at order level.",
    }
    return descriptions.get(table_name, "Derived analytics asset.")


def sample_value(series: pd.Series):
    vals = series.dropna()
    if vals.empty:
        return None
    return str(vals.iloc[0])[:100]


def debug_print_directory(label: str, directory: Path) -> None:
    print(f"[{label}] {directory.resolve()}")
    if directory.exists() and directory.is_dir():
        try:
            files = sorted([p.name for p in directory.iterdir() if p.is_file()])
            print(f"  Files found ({len(files)}): {files}")
        except Exception as exc:
            print(f"  Could not list files: {exc}")
    else:
        print("  Directory does not exist")


def validate_file_map(directory: Path, file_map: Dict[str, str]) -> Tuple[bool, List[str]]:
    missing = []
    for logical_name, file_name in file_map.items():
        if logical_name == "order_reviews":
            continue  # optional
        if not (directory / file_name).exists():
            missing.append(file_name)
    return len(missing) == 0, missing


def resolve_data_directory(explicit_data_dir: Optional[str]) -> Tuple[Path, Dict[str, str]]:
    script_dir = Path(__file__).resolve().parent
    cwd = Path.cwd().resolve()

    candidates: List[Path] = []
    if explicit_data_dir:
        explicit_path = Path(explicit_data_dir)
        if not explicit_path.is_absolute():
            candidates.append((cwd / explicit_path).resolve())
            candidates.append((script_dir / explicit_path).resolve())
        candidates.append(explicit_path.resolve())

    candidates.extend([
        (script_dir / "data").resolve(),
        script_dir.resolve(),
        (cwd / "data").resolve(),
        cwd.resolve(),
    ])

    seen = set()
    unique_candidates = []
    for c in candidates:
        key = str(c)
        if key not in seen:
            seen.add(key)
            unique_candidates.append(c)

    print("\n--- Searching for input CSV files ---")
    for candidate in unique_candidates:
        debug_print_directory("Checking", candidate)
        for file_map in FILE_MAP_VARIANTS:
            ok, missing = validate_file_map(candidate, file_map)
            if ok:
                print(f"\nUsing data directory: {candidate}")
                print(f"Using file map: {file_map}\n")
                return candidate, file_map
            else:
                print(f"  Tried file map {file_map} -> missing required files: {missing}")

    raise FileNotFoundError(
        "Could not find the required input files in any searched location.\n"
        "Tried searching current working directory, script directory, and their ./data subfolders.\n"
        "You can also pass --data-dir explicitly to point to the folder containing the CSV files."
    )


def build_lineage_edges(loaded: Dict[str, pd.DataFrame]) -> pd.DataFrame:
    edges = []
    if "customers" in loaded:
        edges.append(["raw.customers", "curated.dim_customer", "Build customer dimension", "daily", "table"])
    if "products" in loaded:
        edges.append(["raw.products", "curated.dim_product", "Build product dimension", "daily", "table"])

    required_fact = {"orders", "order_items", "payments"}
    if required_fact.issubset(set(loaded.keys())):
        edges.extend([
            ["raw.orders", "curated.fact_order", "Join orders to line items and payments", "daily", "table"],
            ["raw.order_items", "curated.fact_order", "Join orders to line items and payments", "daily", "table"],
            ["raw.payments", "curated.fact_order", "Aggregate payment info into fact", "daily", "table"],
        ])
    if "customers" in loaded and required_fact.issubset(set(loaded.keys())):
        edges.append(["raw.customers", "curated.fact_order", "Enrich fact with customer key", "daily", "table"])
    if "products" in loaded and required_fact.issubset(set(loaded.keys())):
        edges.append(["raw.products", "curated.fact_order", "Enrich fact with product key", "daily", "table"])

    if required_fact.issubset(set(loaded.keys())):
        edges.extend([
            ["curated.fact_order", "mart.sales_summary_daily", "Aggregate daily sales KPIs", "daily", "table"],
            ["curated.dim_customer", "mart.sales_summary_daily", "Support slicing by customer geography", "daily", "table"],
            ["curated.dim_product", "mart.sales_summary_daily", "Support slicing by product category", "daily", "table"],
        ])

    return pd.DataFrame(edges, columns=["source_asset", "target_asset", "process_name", "schedule", "lineage_level"])


def build_column_lineage() -> pd.DataFrame:
    mappings = [
        ["raw.customers.customer_id", "curated.dim_customer.customer_id", "direct"],
        ["raw.customers.customer_unique_id", "curated.dim_customer.customer_unique_id", "direct"],
        ["raw.customers.customer_zip_code_prefix", "curated.dim_customer.customer_zip_code_prefix", "direct"],
        ["raw.customers.customer_city", "curated.dim_customer.customer_city", "direct"],
        ["raw.customers.customer_state", "curated.dim_customer.customer_state", "direct"],
        ["raw.products.product_id", "curated.dim_product.product_id", "direct"],
        ["raw.products.product_category_name", "curated.dim_product.product_category_name", "direct"],
        ["raw.orders.order_id", "curated.fact_order.order_id", "direct"],
        ["raw.orders.customer_id", "curated.fact_order.customer_id", "direct"],
        ["raw.orders.order_purchase_timestamp", "curated.fact_order.order_purchase_timestamp", "direct"],
        ["raw.order_items.order_item_id", "curated.fact_order.order_item_id", "direct"],
        ["raw.order_items.product_id", "curated.fact_order.product_id", "direct"],
        ["raw.order_items.price", "curated.fact_order.item_price", "rename"],
        ["raw.order_items.freight_value", "curated.fact_order.freight_value", "direct"],
        ["raw.payments.payment_type", "curated.fact_order.payment_type", "direct"],
        ["raw.payments.payment_installments", "curated.fact_order.payment_installments", "direct"],
        ["raw.payments.payment_value", "curated.fact_order.payment_value", "direct"],
        ["curated.fact_order.order_purchase_timestamp", "mart.sales_summary_daily.order_purchase_date", "cast(date)"],
        ["curated.fact_order.item_price", "mart.sales_summary_daily.total_item_revenue", "sum"],
        ["curated.fact_order.freight_value", "mart.sales_summary_daily.total_freight_revenue", "sum"],
        ["curated.fact_order.order_id", "mart.sales_summary_daily.order_count", "count_distinct"],
    ]
    return pd.DataFrame(mappings, columns=["source_column", "target_column", "transformation_logic"])


def profile_tables(loaded: Dict[str, pd.DataFrame]) -> pd.DataFrame:
    rows = []
    for table_name, df in loaded.items():
        candidate_keys = detect_candidate_keys(df)
        pk = PK_MAP.get(table_name, [])
        has_pk = all(c in df.columns for c in pk) if pk else False
        rows.append({
            "layer": "raw",
            "table_name": table_name,
            "qualified_name": f"raw.{table_name}",
            "domain": DOMAIN_MAP.get(table_name, "General"),
            "owner": OWNER_MAP.get(table_name, "Unassigned"),
            "update_frequency": UPDATE_FREQUENCY.get(table_name, "Unknown"),
            "row_count": int(len(df)),
            "column_count": int(df.shape[1]),
            "primary_key_defined": ", ".join(pk) if pk else "",
            "primary_key_present": bool(has_pk),
            "candidate_keys_detected": ", ".join(candidate_keys),
            "foreign_keys_defined": "; ".join([f"{c}->{rt}.{rc}" for c, rt, rc in FK_MAP.get(table_name, [])]),
            "contains_pii": contains_pii(df),
            "criticality": infer_criticality(table_name),
            "retention_policy": infer_retention(table_name),
            "description": infer_table_description(table_name),
        })

    derived_assets = [
        ("curated", "dim_customer", "Customer", "Conformed customer dimension for analytics"),
        ("curated", "dim_product", "Product", "Conformed product dimension for analytics"),
        ("curated", "fact_order", "Sales", "Analytics fact table at order-item grain with payment enrichment"),
        ("mart", "sales_summary_daily", "Sales", "Daily sales KPI mart for BI consumption"),
    ]
    for layer, table_name, domain, desc in derived_assets:
        rows.append({
            "layer": layer,
            "table_name": table_name,
            "qualified_name": f"{layer}.{table_name}",
            "domain": domain,
            "owner": OWNER_MAP.get(table_name, "Unassigned"),
            "update_frequency": UPDATE_FREQUENCY.get(table_name, "Unknown"),
            "row_count": np.nan,
            "column_count": np.nan,
            "primary_key_defined": ", ".join(PK_MAP.get(table_name, [])),
            "primary_key_present": np.nan,
            "candidate_keys_detected": "",
            "foreign_keys_defined": "; ".join([f"{c}->{rt}.{rc}" for c, rt, rc in FK_MAP.get(table_name, [])]),
            "contains_pii": table_name in {"dim_customer", "fact_order"},
            "criticality": infer_criticality(table_name),
            "retention_policy": infer_retention(table_name),
            "description": desc,
        })
    return pd.DataFrame(rows)


def profile_columns(loaded: Dict[str, pd.DataFrame]) -> pd.DataFrame:
    rows = []
    for table_name, df in loaded.items():
        pk_set = set(PK_MAP.get(table_name, []))
        fk_lookup = {col: (ref_table, ref_col) for col, ref_table, ref_col in FK_MAP.get(table_name, [])}
        for col in df.columns:
            classification, semantic_role = classify_column(col)
            null_pct = round(df[col].isna().mean() * 100, 2) if len(df) else 0
            distinct_count = int(df[col].nunique(dropna=True))
            rows.append({
                "table_name": table_name,
                "qualified_table_name": f"raw.{table_name}",
                "column_name": col,
                "dtype": str(df[col].dtype),
                "dtype_category": dtype_category(df[col]),
                "nullable": bool(df[col].isna().any()),
                "null_pct": null_pct,
                "distinct_count": distinct_count,
                "is_primary_key": col in pk_set,
                "is_foreign_key": col in fk_lookup,
                "references": f"raw.{fk_lookup[col][0]}.{fk_lookup[col][1]}" if col in fk_lookup else "",
                "data_classification": classification,
                "semantic_role": semantic_role,
                "sample_value": sample_value(df[col]),
            })
    return pd.DataFrame(rows)


def data_quality_checks(loaded: Dict[str, pd.DataFrame]) -> pd.DataFrame:
    checks = []
    def add_check(table_name: str, check_name: str, status: str, details: str, severity: str):
        checks.append({
            "table_name": table_name,
            "check_name": check_name,
            "status": status,
            "severity": severity,
            "details": details,
        })

    for table_name, df in loaded.items():
        add_check(table_name, "row_count_nonzero", "PASS" if len(df) > 0 else "FAIL", f"row_count={len(df)}", "High")
        pk_cols = PK_MAP.get(table_name, [])
        if pk_cols:
            missing_pk_cols = [c for c in pk_cols if c not in df.columns]
            if missing_pk_cols:
                add_check(table_name, "primary_key_columns_present", "FAIL", f"missing_columns={missing_pk_cols}", "High")
            else:
                dup_count = int(df.duplicated(subset=pk_cols).sum())
                add_check(table_name, "primary_key_uniqueness", "PASS" if dup_count == 0 else "FAIL", f"duplicate_rows={dup_count} on {pk_cols}", "High")
                null_pk_rows = int(df[pk_cols].isna().any(axis=1).sum())
                add_check(table_name, "primary_key_not_null", "PASS" if null_pk_rows == 0 else "FAIL", f"null_pk_rows={null_pk_rows} on {pk_cols}", "High")

        for col, ref_table, ref_col in FK_MAP.get(table_name, []):
            if col in df.columns and ref_table in loaded and ref_col in loaded[ref_table].columns:
                violations = int((~df[col].dropna().isin(set(loaded[ref_table][ref_col].dropna()))).sum())
                add_check(table_name, f"fk_integrity_{col}", "PASS" if violations == 0 else "WARN", f"orphan_rows={violations}; references raw.{ref_table}.{ref_col}", "Medium")

        for col in df.columns:
            null_pct = df[col].isna().mean() * 100 if len(df) else 0
            if null_pct > 30:
                add_check(table_name, f"high_null_rate_{col}", "WARN", f"null_pct={null_pct:.2f}%", "Low")
    return pd.DataFrame(checks)


def process_registry() -> pd.DataFrame:
    rows = [
        {
            "process_name": "build_dim_customer",
            "layer": "curated",
            "target_asset": "curated.dim_customer",
            "owners": "Analytics Engineering",
            "schedule": "Daily",
            "upstream_assets": "raw.customers",
            "logic_summary": "Select customer identifiers and geography fields for conformed dimension.",
        },
        {
            "process_name": "build_dim_product",
            "layer": "curated",
            "target_asset": "curated.dim_product",
            "owners": "Analytics Engineering",
            "schedule": "Daily",
            "upstream_assets": "raw.products",
            "logic_summary": "Select product identifiers and descriptive attributes for conformed dimension.",
        },
        {
            "process_name": "build_fact_order",
            "layer": "curated",
            "target_asset": "curated.fact_order",
            "owners": "Analytics Engineering",
            "schedule": "Daily",
            "upstream_assets": "raw.orders, raw.order_items, raw.payments, raw.customers, raw.products",
            "logic_summary": "Join order headers, line items, payments, customer, and product data at order-item grain.",
        },
        {
            "process_name": "build_sales_summary_daily",
            "layer": "mart",
            "target_asset": "mart.sales_summary_daily",
            "owners": "BI / Analytics",
            "schedule": "Daily",
            "upstream_assets": "curated.fact_order, curated.dim_customer, curated.dim_product",
            "logic_summary": "Aggregate daily revenue, freight, and order counts for dashboard consumption.",
        },
    ]
    return pd.DataFrame(rows)


def create_lineage_graph(lineage_df: pd.DataFrame, out_path: Path) -> None:
    if nx is None:
        return
    G = nx.DiGraph()
    for _, row in lineage_df.iterrows():
        G.add_edge(row["source_asset"], row["target_asset"], process=row["process_name"])

    plt.figure(figsize=(14, 8))
    pos = nx.spring_layout(G, seed=42, k=1.2)
    node_colors = []
    for node in G.nodes():
        if node.startswith("raw."):
            node_colors.append("#7aa6c2")
        elif node.startswith("curated."):
            node_colors.append("#7fc97f")
        else:
            node_colors.append("#fdc086")
    nx.draw_networkx_nodes(G, pos, node_size=2200, node_color=node_colors)
    nx.draw_networkx_labels(G, pos, font_size=9)
    nx.draw_networkx_edges(G, pos, arrows=True, arrowstyle='-|>', arrowsize=18, width=1.6)
    edge_labels = {(u, v): d.get("process", "") for u, v, d in G.edges(data=True)}
    nx.draw_networkx_edge_labels(G, pos, edge_labels=edge_labels, font_size=7)
    plt.title("Data Lineage Graph")
    plt.axis("off")
    plt.tight_layout()
    plt.savefig(out_path, dpi=220, bbox_inches="tight")
    plt.close()


def pii_summary(column_catalog: pd.DataFrame) -> pd.DataFrame:
    pii = column_catalog[column_catalog["data_classification"].isin(["Restricted", "Confidential"])].copy()
    if pii.empty:
        return pd.DataFrame(columns=["qualified_table_name", "restricted_columns", "confidential_columns", "total_sensitive_columns"])
    rows = []
    for table_name, g in pii.groupby("qualified_table_name"):
        rows.append({
            "qualified_table_name": table_name,
            "restricted_columns": int((g["data_classification"] == "Restricted").sum()),
            "confidential_columns": int((g["data_classification"] == "Confidential").sum()),
            "total_sensitive_columns": int(len(g)),
        })
    return pd.DataFrame(rows)


def materialize_sample_curated_assets(loaded: Dict[str, pd.DataFrame], output_dir: Path) -> None:
    if "customers" in loaded:
        cust_cols = [c for c in ["customer_id", "customer_unique_id", "customer_zip_code_prefix", "customer_city", "customer_state"] if c in loaded["customers"].columns]
        if cust_cols:
            loaded["customers"][cust_cols].drop_duplicates().to_csv(output_dir / "dim_customer.csv", index=False)

    if "products" in loaded:
        prod_cols = [c for c in ["product_id", "product_category_name", "product_name_lenght", "product_description_lenght", "product_weight_g", "product_length_cm", "product_height_cm", "product_width_cm"] if c in loaded["products"].columns]
        if prod_cols:
            loaded["products"][prod_cols].drop_duplicates().to_csv(output_dir / "dim_product.csv", index=False)

    if {"orders", "order_items", "payments"}.issubset(loaded.keys()):
        orders = loaded["orders"].copy()
        items = loaded["order_items"].copy()
        payments = loaded["payments"].copy()

        if "order_purchase_timestamp" in orders.columns:
            orders["order_purchase_timestamp"] = pd.to_datetime(orders["order_purchase_timestamp"], errors="coerce")

        payment_group = payments.groupby("order_id", dropna=False)
        payment_agg_data = {}
        if "payment_value" in payments.columns:
            payment_agg_data["payment_value"] = ("payment_value", "sum")
        if "payment_installments" in payments.columns:
            payment_agg_data["payment_installments"] = ("payment_installments", "max")
        if "payment_type" in payments.columns:
            payment_agg_data["payment_type"] = ("payment_type", lambda x: x.dropna().astype(str).iloc[0] if len(x.dropna()) else None)
        payment_agg = payment_group.agg(**payment_agg_data).reset_index() if payment_agg_data else payments[["order_id"]].drop_duplicates()

        fact = items.merge(orders, on="order_id", how="left")
        fact = fact.merge(payment_agg, on="order_id", how="left")
        fact.to_csv(output_dir / "fact_order.csv", index=False)

        if "order_purchase_timestamp" in fact.columns:
            fact["order_purchase_date"] = pd.to_datetime(fact["order_purchase_timestamp"], errors="coerce").dt.date
            agg_spec = {"order_id": pd.Series.nunique}
            named_agg = {
                "order_count": ("order_id", pd.Series.nunique),
            }
            if "price" in fact.columns:
                named_agg["total_item_revenue"] = ("price", "sum")
            if "freight_value" in fact.columns:
                named_agg["total_freight_revenue"] = ("freight_value", "sum")
            daily = fact.groupby("order_purchase_date", dropna=False).agg(**named_agg).reset_index()
            daily.to_csv(output_dir / "sales_summary_daily.csv", index=False)


def autosize_worksheet(ws, df: pd.DataFrame) -> None:
    # Freeze header row and style it
    ws.freeze_panes = "A2"
    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for idx, col_name in enumerate(df.columns, start=1):
        series_as_str = df[col_name].astype(str).replace("nan", "") if len(df) else pd.Series([col_name])
        max_len = max([len(str(col_name))] + [len(x) for x in series_as_str.head(500).tolist()])
        ws.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 12), 40)


def safe_sheet_name(file_name: str) -> str:
    base = CSV_SHEET_NAMES.get(file_name, Path(file_name).stem)
    invalid = ['[', ']', '*', '?', '/', '\\', ':']
    for ch in invalid:
        base = base.replace(ch, '_')
    return base[:31]


def add_image_sheet(xlsx_path: Path, image_path: Path) -> None:
    if not image_path.exists():
        return

    wb = load_workbook(xlsx_path)
    if "Lineage_Graph" in wb.sheetnames:
        del wb["Lineage_Graph"]
    ws = wb.create_sheet("Lineage_Graph")

    # Create a readable canvas with no competing worksheet content.
    ws.sheet_view.showGridLines = False
    ws["B2"] = "Data Lineage Graph"
    ws["B2"].font = Font(bold=True, size=14)
    ws["B2"].alignment = Alignment(horizontal="left")
    ws["B3"].alignment = Alignment(wrap_text=True)

    for col in range(2, 22):  # B:V
        ws.column_dimensions[get_column_letter(col)].width = 14
    for row in range(4, 45):
        ws.row_dimensions[row].height = 28

    img = XLImage(str(image_path))

    # Resize image proportionally so it fits well on the dedicated sheet.
    max_width = 1400
    max_height = 900
    orig_w, orig_h = img.width, img.height
    if PILImage is not None:
        try:
            with PILImage.open(image_path) as pil_img:
                orig_w, orig_h = pil_img.size
        except Exception:
            pass

    scale = min(max_width / orig_w, max_height / orig_h, 1.0)
    img.width = int(orig_w * scale)
    img.height = int(orig_h * scale)

    # Anchor image below title area to avoid overlap.
    ws.add_image(img, "B5")
    wb.save(xlsx_path)


def export_excel_package(output_dir: Path, workbook_name: str = "governance_package.xlsx") -> Path:
    xlsx_path = output_dir / workbook_name

    csv_files = sorted(output_dir.glob("*.csv"))
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        # Cover sheet
        cover = pd.DataFrame({
            "Artifact Type": ["CSV outputs", "Lineage image", "Workbook Path"],
            "Value": [len(csv_files), str((output_dir / 'lineage_graph.png').exists()), str(xlsx_path)]
        })
        cover.to_excel(writer, index=False, sheet_name="Overview")
        ws_cover = writer.book["Overview"]
        ws_cover["A1"].font = Font(bold=True)
        ws_cover["B1"].font = Font(bold=True)
        ws_cover.column_dimensions["A"].width = 18
        ws_cover.column_dimensions["B"].width = 100

        for csv_path in csv_files:
            try:
                df = pd.read_csv(csv_path)
            except Exception:
                # Fallback to an error page if something odd happens
                df = pd.DataFrame({"error": [f"Could not read {csv_path.name}"]})
            sheet_name = safe_sheet_name(csv_path.name)
            df.to_excel(writer, index=False, sheet_name=sheet_name)
            ws = writer.book[sheet_name]
            autosize_worksheet(ws, df)

    add_image_sheet(xlsx_path, output_dir / "lineage_graph.png")
    return xlsx_path


def main(data_dir: Optional[str], output_dir: Optional[str]) -> None:
    script_dir = Path(__file__).resolve().parent
    resolved_data_path, selected_file_map = resolve_data_directory(data_dir)

    if output_dir:
        out_path = Path(output_dir)
        if not out_path.is_absolute():
            out_path = (script_dir / out_path).resolve()
    else:
        out_path = (script_dir / "governance_output").resolve()
    out_path.mkdir(parents=True, exist_ok=True)

    loaded: Dict[str, pd.DataFrame] = {}
    for logical_name, file_name in selected_file_map.items():
        file_path = resolved_data_path / file_name
        if file_path.exists():
            loaded[logical_name] = safe_read_csv(file_path)
        elif logical_name != "order_reviews":
            raise FileNotFoundError(f"Required file vanished during load: {file_path}")

    table_catalog = profile_tables(loaded)
    column_catalog = profile_columns(loaded)
    quality_report = data_quality_checks(loaded)
    lineage_edges = build_lineage_edges(loaded)
    column_lineage = build_column_lineage()
    proc_registry = process_registry()
    pii_report = pii_summary(column_catalog)

    table_catalog.to_csv(out_path / "data_catalog.csv", index=False)
    column_catalog.to_csv(out_path / "column_catalog.csv", index=False)
    quality_report.to_csv(out_path / "data_quality_report.csv", index=False)
    lineage_edges.to_csv(out_path / "lineage_edges.csv", index=False)
    column_lineage.to_csv(out_path / "column_lineage.csv", index=False)
    proc_registry.to_csv(out_path / "process_registry.csv", index=False)
    pii_report.to_csv(out_path / "pii_summary.csv", index=False)

    materialize_sample_curated_assets(loaded, out_path)
    create_lineage_graph(lineage_edges, out_path / "lineage_graph.png")

    summary = pd.DataFrame([
        {"artifact": "data_catalog.csv", "rows": len(table_catalog)},
        {"artifact": "column_catalog.csv", "rows": len(column_catalog)},
        {"artifact": "data_quality_report.csv", "rows": len(quality_report)},
        {"artifact": "lineage_edges.csv", "rows": len(lineage_edges)},
        {"artifact": "column_lineage.csv", "rows": len(column_lineage)},
        {"artifact": "process_registry.csv", "rows": len(proc_registry)},
        {"artifact": "pii_summary.csv", "rows": len(pii_report)},
    ])
    summary.to_csv(out_path / "run_summary.csv", index=False)

    excel_package = export_excel_package(out_path, workbook_name="governance_package.xlsx")

    print("\nData governance and lineage artifacts created successfully.")
    print(f"Input folder : {resolved_data_path.resolve()}")
    print(f"Output folder: {out_path.resolve()}")
    print(f"Excel file   : {excel_package.resolve()}\n")
    print(summary.to_string(index=False))


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Build governance and lineage artifacts from multi-table e-commerce CSVs and package them into Excel.")
    parser.add_argument(
        "--data-dir",
        default=None,
        help="Optional folder containing source CSV files. If omitted, the script auto-detects ./data, script folder, and current working directory.",
    )
    parser.add_argument(
        "--output-dir",
        default=None,
        help="Optional folder to write governance artifacts. If omitted, defaults to ./governance_output next to the script.",
    )
    args = parser.parse_args()
    main(args.data_dir, args.output_dir)
